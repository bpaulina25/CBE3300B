import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from potentiostat import Potentiostat
from scipy.signal import savgol_filter
import scipy.stats as stats
import serial
import time
import os
from openpyxl import load_workbook
from serial.tools import list_ports

def get_port():
    
    rodeo_port = None
    arduino_port = None

    ports = list_ports.comports()
    for port in ports:
        hwid = port.hwid.lower()
        desc = port.description.lower()

        if "239a:802b" in hwid:
            rodeo_port = port.device
        elif "2341:0266" in hwid or "arduino" in desc or "giga" in desc:
            arduino_port = port.device

    if not rodeo_port:
        print("Rodeostat not detected.")
    if not arduino_port:
        print("Arduino GIGA not detected.")

    print("Port connection successful")
    for port in ports:
        print(f"  {port.device} - {port.description} - {port.hwid}")

    return rodeo_port, arduino_port

def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    
    if not os.path.exists(filename):
        # Create file with headers
        df.to_excel(filename, sheet_name=sheet_name, index=False)
    else:
        # Open workbook to get the last row
        book = load_workbook(filename)

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            startrow = sheet.max_row
        else:
            startrow = 0  # new sheet

        # Use mode='a' to append, and no need to set writer.book or writer.sheets anymore
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)

def save_reading(value, filename="Blood Sugar History.xlsx", sheet_name="Log"):
    # Get current date and time
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%I:%M")
    am_pm = now.strftime("%p")

    # Create DataFrame with one row
    df = pd.DataFrame([[f"{value:.2f}", date_str, time_str, am_pm]],
                      columns=["Glucose (mg/dL)", "Date", "Time", "AM/PM"])
    
    # Append to Excel
    append_df_to_excel(filename, df, sheet_name)

def run_amperometry(dev, name):
    # Run chronoamperometry test
    t_all, volt_all, curr_all = dev.run_test(name, display='pbar')

    return t_all, volt_all, curr_all

def run_calibration(t_all, volt_all, curr_all):    
    # Constants from the piecewise model
    slope_cutoff = 8.393017421908771
    m1 = 14.6689445644037
    b1 = -11.681609392359228
    m2 = 51.48481214649696
    b2 = -320.6778274115537

    # Smoothing parameters
    window = 50
    smooth_window = 7
    poly = 2 

    # Sort the data from the CA test
    volt_all = np.array(volt_all)  # V
    curr_all = np.array(curr_all)  # μA
    t_all = np.array(t_all)        # s

    # Get current values once step is applied
    mask = volt_all != 0
    t = t_all[5:5+window]
    curr = curr_all[mask][5:5+window]

    # Smooth data
    curr_smooth = savgol_filter(curr, window_length=smooth_window, polyorder=poly)
    t_axis = 1 / np.sqrt(t)
    
    # Perform linear regression
    slope, _, _, _, _ = stats.linregress(t_axis, curr_smooth)
    
    # Convert slope to concentration
    if slope <= slope_cutoff:
        glucose_conc = m1 * slope + b1
    else:
        glucose_conc = m2 * slope + b2
    
    return glucose_conc

def send_to_arduino(value, port, baudrate=115200):
    try:
        with serial.Serial(port, baudrate, timeout=2) as ser:
            time.sleep(2)  # Wait for Arduino to reset
            ser.write((str(value) + '\n').encode())
            print(f"Sent to Arduino: {value}")
    except serial.SerialException as e:
        print(f"Serial error: {e}")

def save_and_send_data(reading, giga_port, 
                              csv_file="Reading_to_GIGA.csv", 
                              excel_file="Blood Sugar History.xlsx", 
                              sheet_name="Log"):
  
    # Round to nearest whole number
    rounded_reading = int(round(reading))

    # Save to CSV for Arduino (currently re-writes same .csv -- change to delete and replace?)
    df_csv = pd.DataFrame([[rounded_reading]])
    df_csv.to_csv(csv_file, index=False, header=False)

    # Save to history log with date/time stamp
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%I:%M")
    am_pm = now.strftime("%p")

    df_log = pd.DataFrame([[rounded_reading, date_str, time_str, am_pm]],
                          columns=["Glucose (mg/dL)", "Date", "Time", "AM/PM"])
    # Append existing log
    append_df_to_excel(excel_file, df_log, sheet_name)

    # Send to Arduino
    send_to_arduino(rounded_reading, port=giga_port)


if __name__ == '__main__':
    
    rodeo_port, giga_port = get_port()
    
    if not rodeo_port:
        print("Please check Rodeostat USB connection and try again.")
        exit()
    
    if not giga_port:
        print("Please check Arduino USB connection and try again.")
        exit()

    #dev = Potentiostat('/dev/tty.usbmodem101')
    dev = Potentiostat(rodeo_port)
    dev.set_curr_range('1000uA')
    dev.set_sample_period(10)
    dev.set_auto_connect(True)

    name = 'chronoamp'
    test_param = {
        'quietValue': 0.0,
        'quietTime': 6000,
        'step': [(1000, 0.200)],
    }
    
    dev.set_param(name, test_param)

# Loop for continuous testing    
while True:
        user_input = input("Add test strip (t) or exit (exit): ").strip().lower()

        if user_input == 'exit':
            print("Exiting program.")
            break
        elif user_input == 't':
            for i in range(5, 0, -1):
                print(f"Running test in {i}...", end='\r')
                time.sleep(1)

            print("Running test...")

            try:
                t_all, volt_all, curr_all = run_amperometry(dev, name)
                reading = run_calibration(t_all, volt_all, curr_all)
                save_and_send_data(reading, giga_port)
                print(f"Glucose: {reading:.0f} mg/dL")
            except Exception as e:
                print(f"Error during test: {str(e)}")

            print("\nTest complete.\n")
        else:
            print("Invalid input. Please type 't' to run a test or 'exit' to quit.\n")
